"""Module for parsing Fluvius grid costs"""
from __future__ import annotations
from io import BytesIO
from urllib.parse import urlsplit, urlunsplit, urljoin

from bs4 import BeautifulSoup, Tag
import requests
from openpyxl import load_workbook

from dao.gridcost import EnergyDirection, EnergyGridCost


def extract_excel_url(url: str) -> str:
    """Get the Excel URL from a given URL"""
    url_details = urlsplit(url)
    base_url = urlunsplit((url_details[0], url_details[1], "", "", ""))
    html_text = requests.get(url).text
    soup = BeautifulSoup(html_text, "html.parser")
    file = soup.find("span", class_="file")
    file_link = file.find("a")
    return urljoin(base_url, file_link["href"])


def iterate_sections(article: Tag):
    """Iterate over the sections"""
    # Find a H2, H3 and DIV combination
    section_name = None
    year = None
    for child in article.children:
        if child.name == "h2":
            section_name = str(child.text).strip()
        if child.name == "h3":
            year = str(child.text).strip()
        if child.name == "div" and section_name is not None and year is not None:
            yield section_name, year, child
            section_name = None
            year = None


def iterate_subsection(article: Tag):
    """Iterate over the subsections"""
    for name, year, section in iterate_sections(article):
        # Find a H4 and DIV combination
        for subsection in section.find_all("div", class_="pg-container__inner"):
            subsection_name = None
            for child in subsection.children:
                if child.name == "h4":
                    subsection_name = str(child.text).strip()
                if child.name == "div" and subsection_name is not None:
                    yield name, year, subsection_name, child
                    subsection_name = None


def iterate_subsection_links(article: Tag, base_url: str):
    """Iterate over the section links"""
    for section_name, year, subsection_name, subsection in iterate_subsection(article):
        for link in subsection.find_all("a", href=True):
            yield section_name, year, subsection_name, link.text, extract_excel_url(urljoin(base_url, link["href"]))


class FluviusParser:
    url: str = "https://www.fluvius.be/nl/factuur-en-tarieven/netkosten"

    @staticmethod
    def from_url() -> list[EnergyGridCost]:
        """Parse the values from URL"""
        url_details = urlsplit(FluviusParser.url)
        base_url = urlunsplit((url_details[0], url_details[1], "", "", ""))
        html_text = requests.get(FluviusParser.url).text
        soup = BeautifulSoup(html_text, "html.parser")

        article = soup.find("article", class_="node--page")
        return list(
            filter(
                lambda item: item is not None,
                [
                    FluviusParser.from_excel(name, subname, provider, link)
                    for name, _year, subname, provider, link in iterate_subsection_links(article, base_url)
                ],
            )
        )

    @staticmethod
    def from_excel(utility: str, direction: str, provider: str, excel_link: str) -> EnergyGridCost:
        """Read the grid costs from excel"""
        direction_enum = EnergyDirection.INJECTION if direction == "Injectie" else EnergyDirection.DRAWDOWN
        if utility == "Elektriciteit" and direction_enum == EnergyDirection.DRAWDOWN:
            response = requests.get(excel_link)
            wb = load_workbook(BytesIO(response.content))
            ws = wb.active
            peak_usage_avg_monthly_cost: float = ws["O15"].value
            peak_usage_kwh: float = ws["O17"].value
            data_management_standard: float = ws["O29"].value
            data_management_dynamic: float = ws["O28"].value
            public_services_kwh: float = ws["O32"].value
            surcharges_kwh: float = ws["O35"].value
            transmission_charges_kwh: float = ws["O37"].value

            return EnergyGridCost(
                "BE",
                provider,
                direction_enum,
                peak_usage_avg_monthly_cost,
                peak_usage_kwh,
                data_management_standard,
                data_management_dynamic,
                public_services_kwh,
                surcharges_kwh,
                transmission_charges_kwh,
            )

        return None
